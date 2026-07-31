#!/usr/bin/env python3
"""Self-test for the universal-file-reader skill."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path


def run(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(cmd, text=True, capture_output=True, check=False)


def require_ok(result: subprocess.CompletedProcess[str], label: str) -> None:
    if result.returncode != 0:
        raise RuntimeError(
            f"{label} failed with exit code {result.returncode}\n"
            f"STDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}"
        )


def main() -> int:
    try:
        from markitdown import MarkItDown  # noqa: F401
        from openpyxl import Workbook
    except ImportError as exc:
        print(
            "SELF-TEST FAIL: missing dependency. Run: python -m pip install -r requirements.txt",
            file=sys.stderr,
        )
        print(str(exc), file=sys.stderr)
        return 2

    scripts_dir = Path(__file__).resolve().parent
    inspect_script = scripts_dir / "inspect_file.py"
    extract_script = scripts_dir / "extract_to_markdown.py"

    with tempfile.TemporaryDirectory(prefix="universal-file-reader-") as tmp:
        root = Path(tmp)

        txt = root / "sample.txt"
        txt.write_text("hello file reader\n", encoding="utf-8")

        csv_path = root / "sample.csv"
        csv_path.write_text("name,value\nalpha,42\n", encoding="utf-8")

        tsv_path = root / "sample.tsv"
        tsv_path.write_text("name\tvalue\nbeta\t7\n", encoding="utf-8")

        json_path = root / "sample.json"
        json_path.write_text(json.dumps({"status": "ok", "n": 3}), encoding="utf-8")

        zip_path = root / "sample.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            archive.writestr("inside.txt", "zip content")

        xlsx_path = root / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Facts"
        ws["A1"] = "value"
        ws["B1"] = 123
        ws["C1"] = "=B1*2"
        wb.save(xlsx_path)

        html_path = root / "sample.html"
        html_path.write_text(
            "<html><body><h1>Universal Reader</h1><p>MarkItDown works.</p></body></html>",
            encoding="utf-8",
        )
        md_path = root / "sample.md"

        cases = [
            (txt, "hello file reader"),
            (csv_path, "alpha"),
            (tsv_path, "beta"),
            (json_path, '"status": "ok"'),
            (zip_path, "inside.txt"),
            (xlsx_path, '"name": "Facts"'),
        ]

        for path, needle in cases:
            result = run([sys.executable, str(inspect_script), str(path)])
            require_ok(result, f"inspect {path.name}")
            if needle not in result.stdout:
                raise RuntimeError(f"inspect {path.name} did not contain expected marker: {needle}")

        result = run(
            [sys.executable, str(extract_script), str(html_path), "-o", str(md_path)]
        )
        require_ok(result, "MarkItDown HTML extraction")
        markdown = md_path.read_text(encoding="utf-8")
        if "Universal Reader" not in markdown or "MarkItDown works" not in markdown:
            raise RuntimeError("MarkItDown output did not contain expected HTML content")

    print("SELF-TEST PASS: universal-file-reader dependencies and core parsers are working.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
