#!/usr/bin/env python3
"""Inspect common local file types with source-fidelity safeguards."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import zipfile
from pathlib import Path
from typing import Any


def _print_json(data: Any) -> None:
    print(json.dumps(data, ensure_ascii=False, indent=2, default=str))


def inspect_excel(path: Path, max_rows: int, max_cols: int) -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Missing dependency: openpyxl. Install requirements.txt for this skill."
        ) from exc

    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    result: dict[str, Any] = {
        "file": str(path),
        "type": "excel",
        "sheets": [],
        "warning": "Formulas are reported as stored; this script does not recalculate workbooks.",
    }

    for ws in workbook.worksheets:
        rows: list[list[dict[str, Any]]] = []
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows)), start=1
        ):
            row_out: list[dict[str, Any]] = []
            for c_idx, cell in enumerate(row[:max_cols], start=1):
                if cell.value is None:
                    continue
                row_out.append(
                    {
                        "coordinate": cell.coordinate,
                        "value": cell.value,
                        "data_type": cell.data_type,
                    }
                )
            if row_out:
                rows.append(row_out)
        result["sheets"].append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "preview": rows,
            }
        )

    _print_json(result)


def inspect_delimited(path: Path, delimiter: str, max_rows: int) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        rows = []
        for idx, row in enumerate(reader):
            if idx >= max_rows:
                break
            rows.append(row)
    _print_json(
        {
            "file": str(path),
            "type": "tsv" if delimiter == "\t" else "csv",
            "delimiter": "\\t" if delimiter == "\t" else delimiter,
            "preview_rows": rows,
        }
    )


def inspect_json(path: Path) -> None:
    with path.open("r", encoding="utf-8-sig") as handle:
        data = json.load(handle)
    _print_json({"file": str(path), "type": "json", "content": data})


def inspect_zip(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as archive:
        members = [
            {
                "name": info.filename,
                "size": info.file_size,
                "compressed_size": info.compress_size,
                "is_dir": info.is_dir(),
            }
            for info in archive.infolist()
        ]
    _print_json({"file": str(path), "type": "zip", "members": members})


def inspect_text(path: Path, max_chars: int) -> None:
    text = path.read_text(encoding="utf-8-sig", errors="strict")
    truncated = len(text) > max_chars
    _print_json(
        {
            "file": str(path),
            "type": "text",
            "truncated": truncated,
            "content": text[:max_chars],
        }
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect a local file without inventing missing data.")
    parser.add_argument("path", type=Path)
    parser.add_argument("--max-rows", type=int, default=100)
    parser.add_argument("--max-cols", type=int, default=50)
    parser.add_argument("--max-chars", type=int, default=200_000)
    args = parser.parse_args()

    path = args.path.resolve()
    if not path.exists() or not path.is_file():
        print(f"UNREADABLE / ERROR: file not found: {path}", file=sys.stderr)
        return 2

    ext = path.suffix.lower()

    try:
        if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            inspect_excel(path, args.max_rows, args.max_cols)
        elif ext == ".csv":
            inspect_delimited(path, ",", args.max_rows)
        elif ext == ".tsv":
            inspect_delimited(path, "\t", args.max_rows)
        elif ext == ".json":
            inspect_json(path)
        elif ext == ".zip":
            inspect_zip(path)
        elif ext in {
            ".txt", ".md", ".log", ".py", ".js", ".ts", ".tsx", ".jsx",
            ".css", ".scss", ".yaml", ".yml", ".toml", ".ini", ".cfg",
            ".sql", ".sh", ".ps1", ".bat", ".c", ".h", ".cpp", ".hpp",
            ".java", ".go", ".rs", ".rb", ".php", ".r",
        }:
            inspect_text(path, args.max_chars)
        else:
            print(
                "UNREADABLE / ERROR: unsupported by inspect_file.py; use extract_to_markdown.py "
                f"for {ext or 'extensionless files'}.",
                file=sys.stderr,
            )
            return 3
    except UnicodeDecodeError as exc:
        print(f"UNREADABLE / ERROR: text encoding could not be decoded as UTF-8: {exc}", file=sys.stderr)
        return 4
    except Exception as exc:  # Deliberately surface parser errors instead of guessing.
        print(f"UNREADABLE / ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 5

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
